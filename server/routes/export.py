import io
import json
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from ..db import execute_sql
from ..config import RESULTS_TABLE, CORRECTIONS_TABLE

router = APIRouter()

# ─── Template rows: (label, field_path, style) ──────────────────────────────
# style: "section" = bold header, no values
#        "total"   = bold row, values from field_path
#        "data"    = normal row, values from field_path (None → 0)
#        "empty"   = blank row
# field_path: dot-path in extracted JSON, None = 0,
#             list of paths = sum of those values
TEMPLATE = [
    # ─── Ativo ────────────────────────────────────────────────────────────────
    ("Ativo",                                         None,                                                              "section"),
    ("Disponibilidades",                              "ativo_circulante.disponibilidades",                               "data"),
    ("Títulos a Receber",                             "ativo_circulante.titulos_a_receber",                              "data"),
    ("Estoques",                                      "ativo_circulante.estoques",                                       "data"),
    ("Adiantamentos",                                 "ativo_circulante.adiantamentos",                                  "data"),
    ("Impostos a Recuperar",                          "ativo_circulante.impostos_a_recuperar",                           "data"),
    ("Outros ativos circulantes",                     "ativo_circulante.outros_ativos_circulantes",                      "data"),
    ("Conta corrente coop/socios/control./Colig.",    "ativo_circulante.conta_corrente_socios_control_colig",            "data"),
    ("Outros ativos financeiros",                     "ativo_circulante.outros_ativos_financeiros",                      "data"),
    ("Ativo Circulante",                              "ativo_circulante.total_ativo_circulante",                         "total"),
    ("Títulos a Receber",                             "ativo_nao_circulante.titulos_a_receber",                          "data"),
    ("Estoques",                                      "ativo_nao_circulante.estoques",                                   "data"),
    ("Adiantamentos",                                 "ativo_nao_circulante.adiantamentos",                              "data"),
    ("Impostos a Recuperar",                          "ativo_nao_circulante.impostos_a_recuperar",                       "data"),
    ("Despesas pagas antecipadamente",                "ativo_nao_circulante.despesas_pagas_antecipadamente",             "data"),
    ("Conta corrente coop/socios/control./Colig.",    "ativo_nao_circulante.conta_corrente_socios_control_colig",        "data"),
    ("Outros realizável a longo prazo",               "ativo_nao_circulante.outros_realizavel_a_longo_prazo",            "data"),
    ("Ativo Não Circulante",                          "ativo_nao_circulante.total_ativo_nao_circulante",                 "total"),
    ("Investimentos",                                 "ativo_permanente.investimentos",                                  "data"),
    ("Imobilizado",                                   "ativo_permanente.imobilizado",                                    "data"),
    ("Intangível / Diferido",                         "ativo_permanente.intangivel_diferido",                            "data"),
    ("Ativo Permanente",                              "ativo_permanente.total_ativo_permanente",                         "total"),
    ("Ativo Total",                                   "ativo_total",                                                     "total"),
    # ─── blank ────────────────────────────────────────────────────────────────
    (None, None, "empty"),
    # ─── Passivo ──────────────────────────────────────────────────────────────
    ("Passivo",                                       None,                                                              "section"),
    ("Fornecedores",                                  "passivo_circulante.fornecedores",                                 "data"),
    ("Financiamentos com instituições de crédito",    "passivo_circulante.financiamentos_com_instituicoes_de_credito",   "data"),
    ("Salários/Contribuições",                        "passivo_circulante.salarios_contribuicoes",                       "data"),
    ("Tributos",                                      "passivo_circulante.tributos",                                     "data"),
    ("Adiantamentos",                                 "passivo_circulante.adiantamentos",                                "data"),
    ("Conta Corrente sócios/coligadas/controladas",   "passivo_circulante.conta_corrente_socios_coligadas_controladas",  "data"),
    ("Outros passivos circulante",                    "passivo_circulante.outros_passivos_circulante",                   "data"),
    ("Provisões",                                     "passivo_circulante.provisoes",                                    "data"),
    ("Outros passivos financeiros",                   "passivo_circulante.outros_passivos_financeiros",                  "data"),
    ("Passivo Circulante",                            "passivo_circulante.total_passivo_circulante",                     "total"),
    ("Fornecedores",                                  "passivo_nao_circulante.fornecedores",                             "data"),
    ("Financiamentos com instituições de crédito",    "passivo_nao_circulante.financiamentos_com_instituicoes_de_credito","data"),
    ("Salários/Contribuições",                        "passivo_nao_circulante.salarios_contribuicoes",                   "data"),
    ("Tributos",                                      "passivo_nao_circulante.tributos",                                 "data"),
    ("Adiantamentos",                                 "passivo_nao_circulante.adiantamentos",                            "data"),
    ("Conta Corrente sócios/coligadas/controladas",   "passivo_nao_circulante.conta_corrente_socios_coligadas_controladas","data"),
    ("Outros Passivos Não Circulantes",               "passivo_nao_circulante.outros_passivos_nao_circulantes",           "data"),
    ("Provisões",                                     "passivo_nao_circulante.provisoes",                                "data"),
    ("Passivo Não Circulante",                        "passivo_nao_circulante.total_passivo_nao_circulante",             "total"),
    ("Capital Social",                                "patrimonio_liquido.capital_social",                               "data"),
    ("Reserva de capital",                            "patrimonio_liquido.reserva_de_capital",                           "data"),
    ("Reservas de lucro",                             "patrimonio_liquido.reservas_de_lucro",                            "data"),
    ("Reservas de reavaliação",                       "patrimonio_liquido.reservas_de_reavaliacao",                      "data"),
    ("Outras reservas",                               "patrimonio_liquido.outras_reservas",                              "data"),
    ("Lucros ou prejuízos acumulados",                "patrimonio_liquido.lucros_ou_prejuizos_acumulados",               "data"),
    ("Ações em tesouraria",                           "patrimonio_liquido.acoes_em_tesouraria",                          "data"),
    ("Patrimônio líquido",                            "patrimonio_liquido.total_patrimonio_liquido",                     "total"),
    ("Passivo Total",                                 "passivo_total",                                                   "total"),
    # ─── blank ────────────────────────────────────────────────────────────────
    (None, None, "empty"),
    # ─── DRE ──────────────────────────────────────────────────────────────────
    ("DRE",                                           None,                                                              "section"),
    ("Receita de venda produto/mercadoria",           "dre.receita_venda_produto_mercadoria",                            "data"),
    ("Receita serviços/arrendamento",                 "dre.receita_servicos_arrendamento",                               "data"),
    ("Receita operacional bruta",                     "dre.receita_operacional_bruta",                                   "total"),
    ("Vendas anuladas",                               "dre.vendas_anuladas",                                             "data"),
    ("Abatimentos",                                   "dre.abatimentos",                                                 "data"),
    ("Impostos incidentes sobre vendas",              "dre.impostos_incidentes_sobre_vendas",                            "data"),
    ("Deduções de receita bruta",                     "dre.total_deducoes",                                              "total"),
    ("Incentivos a exportações",                      "dre.incentivos_a_exportacoes",                                    "data"),
    ("Incentivos a exportações",                      "dre.incentivos_a_exportacoes",                                    "total"),
    ("Receita operacional líquida",                   "dre.receita_operacional_liquida",                                 "total"),
    ("Custo Serviços/Produtos/Mercadorias Vendidas",  "dre.custo_servicos_produtos_mercadorias_vendidas",                "data"),
    ("Superveniências ativas",                        "dre.superveniencias_ativas",                                      "data"),
    ("Custo",                                         "dre.total_custo",                                                 "total"),
    ("Lucro Bruto",                                   "dre.lucro_bruto",                                                 "total"),
    ("Despesas com vendas",                           "dre.despesas_com_vendas",                                         "data"),
    ("Provisão para devedores duvidosos",             "dre.provisao_para_devedores_duvidosos",                           "data"),
    ("Outras Receitas(-)/Despesas Operacionais",      "dre.outras_receitas_despesas_operacionais",                       "data"),
    ("Despesas administrativas",                      "dre.despesas_administrativas",                                    "data"),
    ("Despesas tributarias",                          "dre.despesas_tributarias",                                        "data"),
    ("Despesas gerais",                               "dre.despesas_gerais",                                             "data"),
    ("Depreciação",                                   "dre.depreciacao",                                                 "data"),
    ("Amortização",                                   "dre.amortizacao",                                                 "data"),
    ("Despesas operacionais",                         "dre.total_despesas_operacionais",                                 "total"),
    ("Lucro operacional",                             "dre.lucro_operacional",                                           "total"),
    ("Encargos Financeiros",                          "dre.encargos_financeiros",                                        "data"),
    ("Descontos concedidos",                          "dre.descontos_concedidos",                                        "data"),
    ("Variação cambial/Corr. Monetária não paga",     "dre.variacao_cambial_nao_paga",                                   "data"),
    ("Despesas financeiras",                          "dre.despesas_financeiras",                                        "total"),
    ("Receitas financeiras",                          "dre.receitas_financeiras",                                        "data"),
    ("Variação cambial/Corr. Monetária não recebida", "dre.variacao_cambial_nao_recebida",                               "data"),
    ("Receitas financeiras",                          "dre.total_receitas_financeiras",                                  "total"),
    ("Lucro Financeiro",                              "dre.lucro_financeiro",                                            "total"),
    ("(-/+)Resultado de equivalência patrimonial",    "dre.resultado_de_equivalencia_patrimonial",                       "data"),
    ("Receita não operacional",                       "dre.receita_nao_operacional",                                     "data"),
    ("(-)Despesa não operacional",                    "dre.despesa_nao_operacional",                                     "data"),
    ("(-/+)Saldo de correção monetária",              "dre.saldo_correcao_monetaria",                                    "data"),
    ("Resultado de alienação de ativos",              "dre.resultado_alienacao_ativos",                                  "data"),
    ("Lucro antes do imposto de renda",               "dre.lucro_antes_imposto_de_renda",                                "total"),
    ("(-)Provisão para imposto de renda",             "dre.provisao_imposto_de_renda",                                   "data"),
    ("(-)CSLL",                                       "dre.csll",                                                        "data"),
    ("Lucro antes das participações",                 "dre.lucro_antes_participacoes",                                   "total"),
    ("Partic. e gratificações estatutárias",          "dre.participacoes_gratificacoes_estatutarias",                    "data"),
    ("Lucro antes da participação minoritária",       "dre.lucro_antes_participacao_minoritaria",                        "total"),
    ("Participação de minoritários",                  "dre.participacao_minoritarios",                                   "data"),
    ("Lucro líquido",                                 "dre.lucro_liquido",                                               "total"),
    ("Dividendos",                                    None,                                                              "data"),
    ("OBSERVAÇÕES",                                   None,                                                              "empty"),
    ("OBSERVAÇÕES FLUXO DE CAIXA",                    None,                                                              "empty"),
    ("Ajuste Fluxo de Caixa (Luc Distrib / Aj Pat / Res Reav)", None,                                                   "data"),
]


def _get(data: dict, path):
    """Get value by dot-path; list of paths → sum."""
    if isinstance(path, list):
        total = 0.0
        for p in path:
            v = _get(data, p)
            if v is not None:
                try:
                    total += float(v)
                except (ValueError, TypeError):
                    pass
        return total
    parts = path.split(".")
    cur = data
    for p in parts:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(p)
    return cur


def _num(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _sheet_name(name: str, used: set) -> str:
    """Safe, unique Excel sheet name (max 31 chars)."""
    safe = name[:28].strip().replace("/", "-").replace("\\", "-").replace("?", "").replace("*", "").replace("[", "").replace("]", "").replace(":", "")
    base, i = safe, 1
    while safe in used:
        safe = f"{base[:27]}{i}"
        i += 1
    used.add(safe)
    return safe


@router.get("/export/excel")
def export_excel(document: str | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Styles ────────────────────────────────────────────────────────────────
    font_bold     = Font(bold=True, size=10)
    font_normal   = Font(size=10)
    font_header   = Font(bold=True, color="FFFFFF", size=10)
    fill_section  = PatternFill("solid", fgColor="D9E1F2")   # light blue
    fill_total    = PatternFill("solid", fgColor="EBF0F7")   # pale blue
    fill_meta     = PatternFill("solid", fgColor="1F3864")   # dark blue (row 1)
    fill_doc_hdr  = PatternFill("solid", fgColor="2E75B6")   # medium blue (rows 2-5)
    fill_corr     = PatternFill("solid", fgColor="FFF3CD")   # yellow (corrected)
    align_center  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
    align_right   = Alignment(horizontal="right",  vertical="center")
    thin          = Side(style="thin", color="BDD7EE")
    border_light  = Border(left=thin, right=thin, top=thin, bottom=thin)

    number_fmt = '#,##0.00'

    def style_cell(cell, bold=False, fill=None, align=None, number=False, border=True):
        cell.font = font_bold if bold else font_normal
        if fill:
            cell.fill = fill
        if align:
            cell.alignment = align
        if number:
            cell.number_format = number_fmt
        if border:
            cell.border = border_light

    # ── Fetch data ────────────────────────────────────────────────────────────
    if document:
        params = [{"name": "doc", "value": document}]
        docs = execute_sql(
            f"SELECT document_name, tipo_entidade, periodo, extracted_json "
            f"FROM {RESULTS_TABLE} WHERE document_name = :doc ORDER BY document_name, periodo, tipo_entidade",
            params,
        )
        corrections_rows = execute_sql(
            f"SELECT document_name, campo, valor_extraido, valor_correto, comentario "
            f"FROM {CORRECTIONS_TABLE} WHERE document_name = :doc ORDER BY document_name, campo",
            params,
        )
    else:
        docs = execute_sql(
            f"SELECT document_name, tipo_entidade, periodo, extracted_json "
            f"FROM {RESULTS_TABLE} ORDER BY document_name, periodo, tipo_entidade"
        )
        corrections_rows = execute_sql(
            f"SELECT document_name, campo, valor_extraido, valor_correto, comentario "
            f"FROM {CORRECTIONS_TABLE} ORDER BY document_name, campo"
        )

    # Build correction index: (doc, campo) → row
    corr_index: dict[tuple, dict] = {}
    for row in corrections_rows:
        corr_index[(row["document_name"], row["campo"])] = row

    # Group records by document_name
    from collections import defaultdict
    doc_records: dict[str, list] = defaultdict(list)
    for row in docs:
        data = row["extracted_json"]
        if isinstance(data, str):
            try:
                data = json.loads(data)
            except Exception:
                data = {}
        doc_records[row["document_name"]].append({
            "tipo_entidade": row.get("tipo_entidade") or "",
            "periodo":       row.get("periodo") or "",
            "data":          data if isinstance(data, dict) else {},
        })

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    used_names: set = set()

    # ── One sheet per document ────────────────────────────────────────────────
    for doc_name, records in doc_records.items():
        # Use razao_social from first record as sheet name
        first_data = records[0]["data"]
        razao = (first_data.get("razao_social") or doc_name)[:28]
        sheet_name = _sheet_name(razao, used_names)
        ws = wb.create_sheet(title=sheet_name)

        # Get moeda/escala from first record
        moeda  = _get(first_data, "identificacao.moeda") or "Real"
        escala = _get(first_data, "identificacao.escala_valores") or "UNIDADE"
        cnpj   = _get(first_data, "cnpj") or ""

        # Data columns: one per (tipo_entidade, periodo) record
        # Column B = labels, columns C+ = data
        n_cols = len(records)
        data_col_start = 3  # C

        # ── Row 1: MOEDA ─────────────────────────────────────────────────────
        ws.cell(1, 1, "MOEDA").font = Font(bold=True, color="FFFFFF", size=10)
        ws.cell(1, 1).fill = fill_meta
        ws.cell(1, 1).alignment = align_center
        moeda_str = f"{moeda} ({escala})"
        ws.cell(1, 2, moeda_str).font = Font(bold=True, color="FFFFFF", size=10)
        ws.cell(1, 2).fill = fill_meta
        ws.cell(1, 2).alignment = align_center
        for ci in range(data_col_start, data_col_start + n_cols):
            ws.cell(1, ci).fill = fill_meta

        # ── Rows 2-5: DOCUMENTO, PERÍODO, TIPO, Nº MESES ─────────────────────
        meta_labels = [
            "DOCUMENTO",
            "PERÍODO",
            "TIPO",
            "Nº MESES DO DEMONSTRATIVO",
        ]
        meta_values = [
            lambda r, d: (d.get("razao_social") or "") + (f" | {cnpj}" if cnpj else ""),
            lambda r, d: r["periodo"],
            lambda r, d: _get(d, "identificacao.tipo_demonstrativo") or "anual",
            lambda r, d: 12,
        ]
        for mi, (lbl, val_fn) in enumerate(zip(meta_labels, meta_values), 2):
            c = ws.cell(mi, 2, lbl)
            c.font = font_header; c.fill = fill_doc_hdr; c.alignment = align_left
            for ci, rec in enumerate(records, data_col_start):
                v = val_fn(rec, rec["data"])
                cell = ws.cell(mi, ci, v)
                cell.font = font_header; cell.fill = fill_doc_hdr
                cell.alignment = align_center

        # ── Row 6: empty ──────────────────────────────────────────────────────
        ws.row_dimensions[6].height = 6

        # ── Template rows starting at row 7 ───────────────────────────────────
        row_num = 7
        for label, field_path, style in TEMPLATE:
            if style == "empty":
                ws.row_dimensions[row_num].height = 8
                row_num += 1
                continue

            # Column A: always empty
            # Column B: label
            lbl_cell = ws.cell(row_num, 2, label)
            is_bold = style in ("section", "total")
            lbl_cell.font = font_bold if is_bold else font_normal
            lbl_cell.alignment = align_left
            if style == "section":
                lbl_cell.fill = fill_section
                lbl_cell.border = border_light
            elif style == "total":
                lbl_cell.fill = fill_total
                lbl_cell.border = border_light
            else:
                lbl_cell.border = border_light

            # Data columns
            if style == "section":
                # Section headers: no data values, just fill
                for ci in range(data_col_start, data_col_start + n_cols):
                    c = ws.cell(row_num, ci)
                    c.fill = fill_section
                    c.border = border_light
            else:
                for ci, rec in enumerate(records, data_col_start):
                    d = rec["data"]
                    # Check if field was corrected
                    corrected = corr_index.get((doc_name, field_path if isinstance(field_path, str) else ""))

                    if field_path is None:
                        val = 0.0
                    else:
                        raw = _get(d, field_path)
                        val = _num(raw)

                    if corrected:
                        try:
                            val = float(corrected["valor_correto"])
                        except (ValueError, TypeError):
                            pass

                    cell = ws.cell(row_num, ci, val)
                    cell.number_format = number_fmt
                    cell.alignment = align_right
                    cell.font = font_bold if is_bold else font_normal
                    if style == "total":
                        cell.fill = fill_total
                    if corrected:
                        cell.fill = fill_corr
                    cell.border = border_light

            row_num += 1

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 2
        ws.column_dimensions["B"].width = 42
        for ci in range(data_col_start, data_col_start + n_cols):
            ws.column_dimensions[get_column_letter(ci)].width = 18

        ws.freeze_panes = f"{get_column_letter(data_col_start)}7"

    # ── Correções sheet ───────────────────────────────────────────────────────
    ws_corr = wb.create_sheet("Correções")
    corr_hdrs = ["Documento", "Campo", "Valor Extraído", "Valor Correto", "Comentário"]
    for c, h in enumerate(corr_hdrs, 1):
        cell = ws_corr.cell(1, c, h)
        cell.font = font_header; cell.fill = PatternFill("solid", fgColor="2E75B6")
        cell.alignment = align_center
        cell.border = border_light
    ws_corr.row_dimensions[1].height = 28
    for r, row in enumerate(corrections_rows, 2):
        for c, key in enumerate(["document_name", "campo", "valor_extraido", "valor_correto", "comentario"], 1):
            cell = ws_corr.cell(r, c, row.get(key, ""))
            cell.alignment = align_left; cell.border = border_light
    ws_corr.column_dimensions["A"].width = 35
    ws_corr.column_dimensions["B"].width = 45
    ws_corr.column_dimensions["C"].width = 18
    ws_corr.column_dimensions["D"].width = 18
    ws_corr.column_dimensions["E"].width = 40

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{document.replace('.pdf', '')}.xlsx" if document else "techfin_resultados.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
