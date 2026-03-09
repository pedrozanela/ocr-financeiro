import io
import json
from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from ..db import execute_sql
from ..config import RESULTS_TABLE, CORRECTIONS_TABLE

router = APIRouter()

# Mirrors fieldDefinitions.ts — (label, dot-path, is_number)
FIELDS = [
    # Identificação
    ("Razão Social",                              "razao_social",                                                           False),
    ("CNPJ",                                      "cnpj",                                                                   False),
    ("Período",                                   "identificacao.periodo",                                                  False),
    ("Tipo de Demonstrativo",                     "identificacao.tipo_demonstrativo",                                       False),
    ("Moeda",                                     "identificacao.moeda",                                                    False),
    ("Escala de Valores",                         "identificacao.escala_valores",                                           False),
    # Ativo Circulante
    ("Disponibilidades",                          "ativo_circulante.disponibilidades",                                      True),
    ("Títulos a Receber (AC)",                    "ativo_circulante.titulos_a_receber",                                     True),
    ("Estoques (AC)",                             "ativo_circulante.estoques",                                              True),
    ("Adiantamentos (AC)",                        "ativo_circulante.adiantamentos",                                         True),
    ("Impostos a Recuperar (CP)",                 "ativo_circulante.impostos_a_recuperar",                                  True),
    ("Outros Ativos Circulantes",                 "ativo_circulante.outros_ativos_circulantes",                             True),
    ("C/C Sócios / Coligadas (AC)",               "ativo_circulante.conta_corrente_socios_control_colig",                  True),
    ("Outros Ativos Financeiros (AC)",            "ativo_circulante.outros_ativos_financeiros",                             True),
    ("Total Ativo Circulante",                    "ativo_circulante.total_ativo_circulante",                                True),
    # Ativo Não Circulante
    ("Títulos a Receber (LP)",                    "ativo_nao_circulante.titulos_a_receber",                                 True),
    ("Estoques (LP)",                             "ativo_nao_circulante.estoques",                                          True),
    ("Adiantamentos (LP)",                        "ativo_nao_circulante.adiantamentos",                                     True),
    ("Impostos a Recuperar (LP)",                 "ativo_nao_circulante.impostos_a_recuperar",                              True),
    ("Despesas Pagas Antecipadamente",            "ativo_nao_circulante.despesas_pagas_antecipadamente",                   True),
    ("C/C Sócios / Coligadas (ANC)",              "ativo_nao_circulante.conta_corrente_socios_control_colig",              True),
    ("Outros Realizável LP",                      "ativo_nao_circulante.outros_realizavel_a_longo_prazo",                  True),
    ("Total Ativo Não Circulante",                "ativo_nao_circulante.total_ativo_nao_circulante",                        True),
    # Ativo Permanente
    ("Investimentos",                             "ativo_permanente.investimentos",                                         True),
    ("Imobilizado",                               "ativo_permanente.imobilizado",                                           True),
    ("Intangível / Diferido",                     "ativo_permanente.intangivel_diferido",                                   True),
    ("Total Ativo Permanente",                    "ativo_permanente.total_ativo_permanente",                                True),
    ("Ativo Total",                               "ativo_total",                                                            True),
    # Passivo Circulante
    ("Fornecedores (CP)",                         "passivo_circulante.fornecedores",                                        True),
    ("Financiamentos (CP)",                       "passivo_circulante.financiamentos_com_instituicoes_de_credito",          True),
    ("Salários e Contribuições (CP)",             "passivo_circulante.salarios_contribuicoes",                              True),
    ("Tributos (CP)",                             "passivo_circulante.tributos",                                            True),
    ("Adiantamentos de Clientes (CP)",            "passivo_circulante.adiantamentos",                                       True),
    ("C/C Sócios / Coligadas (PC)",               "passivo_circulante.conta_corrente_socios_coligadas_controladas",        True),
    ("Outros Passivos Circulantes",               "passivo_circulante.outros_passivos_circulante",                          True),
    ("Provisões (CP)",                            "passivo_circulante.provisoes",                                           True),
    ("Outros Passivos Financeiros (CP)",          "passivo_circulante.outros_passivos_financeiros",                         True),
    ("Total Passivo Circulante",                  "passivo_circulante.total_passivo_circulante",                            True),
    # Passivo Não Circulante
    ("Fornecedores (LP)",                         "passivo_nao_circulante.fornecedores",                                    True),
    ("Financiamentos (LP)",                       "passivo_nao_circulante.financiamentos_com_instituicoes_de_credito",      True),
    ("Salários e Contribuições (LP)",             "passivo_nao_circulante.salarios_contribuicoes",                          True),
    ("Tributos (LP)",                             "passivo_nao_circulante.tributos",                                        True),
    ("Adiantamentos de Clientes (LP)",            "passivo_nao_circulante.adiantamentos",                                   True),
    ("C/C Sócios / Coligadas (PNC)",              "passivo_nao_circulante.conta_corrente_socios_coligadas_controladas",    True),
    ("Outros Passivos Não Circulantes",           "passivo_nao_circulante.outros_passivos_nao_circulantes",                True),
    ("Provisões (LP)",                            "passivo_nao_circulante.provisoes",                                       True),
    ("Total Passivo Não Circulante",              "passivo_nao_circulante.total_passivo_nao_circulante",                    True),
    # Patrimônio Líquido
    ("Capital Social",                            "patrimonio_liquido.capital_social",                                      True),
    ("Reserva de Capital",                        "patrimonio_liquido.reserva_de_capital",                                  True),
    ("Reservas de Lucro",                         "patrimonio_liquido.reservas_de_lucro",                                   True),
    ("Reservas de Reavaliação",                   "patrimonio_liquido.reservas_de_reavaliacao",                             True),
    ("Outras Reservas",                           "patrimonio_liquido.outras_reservas",                                     True),
    ("Lucros / Prejuízos Acumulados",             "patrimonio_liquido.lucros_ou_prejuizos_acumulados",                      True),
    ("Ações em Tesouraria",                       "patrimonio_liquido.acoes_em_tesouraria",                                 True),
    ("Total Patrimônio Líquido",                  "patrimonio_liquido.total_patrimonio_liquido",                            True),
    ("Passivo Total",                             "passivo_total",                                                          True),
    # DRE
    ("Receita Operacional Bruta",                 "dre.receita_operacional_bruta",                                          True),
    ("Vendas Anuladas",                           "dre.vendas_anuladas",                                                    True),
    ("Abatimentos",                               "dre.abatimentos",                                                        True),
    ("Impostos sobre Vendas",                     "dre.impostos_incidentes_sobre_vendas",                                   True),
    ("Total Deduções",                            "dre.total_deducoes",                                                     True),
    ("Receita Operacional Líquida",               "dre.receita_operacional_liquida",                                        True),
    ("CMV / CPV / CSP",                           "dre.custo_servicos_produtos_mercadorias_vendidas",                       True),
    ("Lucro Bruto",                               "dre.lucro_bruto",                                                        True),
    ("Despesas com Vendas",                       "dre.despesas_com_vendas",                                                True),
    ("Provisão p/ Dev. Duvidosos",                "dre.provisao_para_devedores_duvidosos",                                  True),
    ("Outras Receitas/Despesas Operacionais",     "dre.outras_receitas_despesas_operacionais",                              True),
    ("Despesas Administrativas",                  "dre.despesas_administrativas",                                           True),
    ("Despesas Tributárias",                      "dre.despesas_tributarias",                                               True),
    ("Despesas Gerais",                           "dre.despesas_gerais",                                                    True),
    ("Depreciação",                               "dre.depreciacao",                                                        True),
    ("Amortização",                               "dre.amortizacao",                                                        True),
    ("Total Despesas Operacionais",               "dre.total_despesas_operacionais",                                        True),
    ("Lucro Operacional (EBIT)",                  "dre.lucro_operacional",                                                  True),
    ("Despesas Financeiras",                      "dre.despesas_financeiras",                                               True),
    ("Receitas Financeiras",                      "dre.receitas_financeiras",                                               True),
    ("Resultado Financeiro",                      "dre.total_resultado_financeiro",                                         True),
    ("Equivalência Patrimonial",                  "dre.resultado_de_equivalencia_patrimonial",                              True),
    ("LAIR",                                      "dre.lucro_antes_imposto_de_renda",                                       True),
    ("Provisão IRPJ / CSLL",                      "dre.provisao_imposto_de_renda_csll",                                     True),
    ("Lucro Líquido",                             "dre.lucro_liquido",                                                      True),
]


def _get(data: dict, path: str):
    parts = path.split(".")
    cur = data
    for p in parts:
        if not isinstance(cur, dict):
            return None
        cur = cur.get(p)
    return cur


def _to_number(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return val


@router.get("/export/excel")
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Fetch data ──────────────────────────────────────────────────────────
    docs = execute_sql(f"SELECT document_name, tipo_entidade, periodo, extracted_json FROM {RESULTS_TABLE} ORDER BY document_name, tipo_entidade, periodo")
    corrections_rows = execute_sql(f"SELECT document_name, campo, valor_extraido, valor_correto, comentario FROM {CORRECTIONS_TABLE} ORDER BY document_name, campo")

    # corrections indexed by (doc, campo)
    corr_index: dict[tuple, dict] = {}
    for row in corrections_rows:
        corr_index[(row["document_name"], row["campo"])] = row

    # ── Workbook ─────────────────────────────────────────────────────────────
    wb = Workbook()

    # ── Styles ───────────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor="0F2137")
    corrected_fill = PatternFill("solid", fgColor="FFF3CD")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Sheet 1: Dados Financeiros ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Dados Financeiros"

    col_headers = ["Documento", "Tipo Entidade", "Período"] + [label for label, _, _ in FIELDS] + ["Correções"]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "D2"

    for r, row in enumerate(docs, 2):
        data = row["extracted_json"]
        if isinstance(data, str):
            data = json.loads(data)

        doc_name = row["document_name"]
        tipo = row.get("tipo_entidade") or ""
        periodo = row.get("periodo") or ""
        n_corr = sum(1 for (d, _) in corr_index if d == doc_name)

        for c, val in enumerate([doc_name, tipo, periodo], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = left
            cell.border = border

        for c, (_, path, is_number) in enumerate(FIELDS, 4):
            raw = _get(data, path)
            corrected = corr_index.get((doc_name, path))

            if is_number:
                val = _to_number(raw)
            else:
                val = str(raw) if raw is not None else ""

            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = left
            cell.border = border
            if corrected:
                cell.fill = corrected_fill
                try:
                    cell.value = float(corrected["valor_correto"]) if is_number else corrected["valor_correto"]
                except (ValueError, TypeError):
                    cell.value = corrected["valor_correto"]

        # correções count
        cell = ws.cell(row=r, column=len(col_headers), value=n_corr if n_corr else "")
        cell.alignment = center
        cell.border = border

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    for c in range(4, len(col_headers)):
        ws.column_dimensions[get_column_letter(c)].width = 18
    ws.column_dimensions[get_column_letter(len(col_headers))].width = 12

    # ── Sheet 2: Correções ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Correções")
    corr_headers = ["Documento", "Campo", "Valor Extraído", "Valor Correto", "Comentário"]
    for c, h in enumerate(corr_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws2.row_dimensions[1].height = 30

    for r, row in enumerate(corrections_rows, 2):
        for c, key in enumerate(["document_name", "campo", "valor_extraido", "valor_correto", "comentario"], 1):
            cell = ws2.cell(row=r, column=c, value=row.get(key, ""))
            cell.alignment = left
            cell.border = border

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 45
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 18
    ws2.column_dimensions["E"].width = 40

    # ── Sheet 3: Fontes ───────────────────────────────────────────────────────
    PATH_TO_LABEL = {path: label for label, path, _ in FIELDS}

    ws3 = wb.create_sheet("Fontes")
    fontes_headers = ["Documento", "Tipo Entidade", "Período", "Campo", "Fonte"]
    for c, h in enumerate(fontes_headers, 1):
        cell = ws3.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws3.row_dimensions[1].height = 30
    ws3.freeze_panes = "A2"

    fonte_row = 2
    for row in docs:
        data = row["extracted_json"]
        if isinstance(data, str):
            data = json.loads(data)
        doc_name = row["document_name"]
        tipo = row.get("tipo_entidade") or ""
        periodo = row.get("periodo") or ""
        fontes: dict = data.get("fontes") if isinstance(data, dict) else {}
        if not fontes:
            continue
        for path, fonte_text in fontes.items():
            label = PATH_TO_LABEL.get(path, path)
            for c, val in enumerate([doc_name, tipo, periodo, label, fonte_text], 1):
                cell = ws3.cell(row=fonte_row, column=c, value=val or "")
                cell.alignment = left
                cell.border = border
            fonte_row += 1

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 14
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 35
    ws3.column_dimensions["E"].width = 80

    # ── Stream response ───────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=techfin_resultados.xlsx"},
    )
